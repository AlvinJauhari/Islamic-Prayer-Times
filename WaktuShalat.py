from mpmath import *
mp.dps = 25; mp.pretty = True
from openpyxl import load_workbook
from tkinter import *
from tkinter import ttk
import tkinter.font

root = Tk()
root.title('WAKTU SHALAT')
root.geometry("1150x600")

changefont = tkinter.font.Font(size=11)
changefont1 = tkinter.font.Font(size=9)
changefont2 = tkinter.font.Font(size=10)

# Tema Tabel
style = ttk.Style()
style.theme_use('clam')

# Scroll Tabel
framescroll = Frame(root)
framescroll.place(x=550,y=40)
scroll = Scrollbar(framescroll)
scroll.pack(side=RIGHT,fill=Y)

tabel = ttk.Treeview(framescroll,height=25,yscrollcommand=scroll.set,selectmode='browse')
tabel.pack()
#scroll.config(command=tabel.yview)
tabel['columns'] = ('Tanggal','Subuh','Terbit Matahari','Zuhur','Ashar','Maghrib','Isya','Zona Waktu')

# Data Tabel
tabel.column('#0',width=0,stretch=NO)
tabel.column('Tanggal',anchor=CENTER,width=50)
tabel.column('Subuh',anchor=CENTER,width=70)
tabel.column('Terbit Matahari',anchor=CENTER,width=70)
tabel.column('Zuhur',anchor=CENTER,width=70)
tabel.column('Ashar',anchor=CENTER,width=70)
tabel.column('Maghrib',anchor=CENTER,width=70)
tabel.column('Isya',anchor=CENTER,width=70)
tabel.column('Zona Waktu',anchor=CENTER,width=70)

# Heading Tabel
tabel.heading('#0',text='')
tabel.heading('Tanggal',text='Tgl',anchor=CENTER)
tabel.heading('Subuh',text='Subuh',anchor=CENTER)
tabel.heading('Terbit Matahari',text='Fajr',anchor=CENTER)
tabel.heading('Zuhur',text='Zuhur',anchor=CENTER)
tabel.heading('Ashar',text='Ashar',anchor=CENTER)
tabel.heading('Maghrib',text='Maghrib',anchor=CENTER)
tabel.heading('Isya',text='Isya',anchor=CENTER)
tabel.heading('Zona Waktu',text='Zona',anchor=CENTER)

# Label & Frame Output
labelkt = Label(root,text='Kota:',font=changefont)
labelkt.place(x=935,y=15)

frame0 = LabelFrame(root,padx=220,pady=85,bg='grey')
frame0.place(x=40,y=325)
label0 = Label(frame0,text='').grid()

frame1 = LabelFrame(root,padx=220,pady=80,bg='#d3d3d3')
frame1.place(x=40,y=370)
label1 = Label(frame1,text='').grid()

frame2 = LabelFrame(root,padx=5,pady=1)
frame2.place(x=80,y=380)
label2 = Label(frame2,text='Subuh',font=changefont2).grid()

frame3 = LabelFrame(root,padx=10,pady=1)
frame3.place(x=230,y=380)
label3 = Label(frame3,text='Fajr',font=changefont2).grid()

frame4 = LabelFrame(root,padx=5,pady=1)
frame4.place(x=378,y=380)
label4 = Label(frame4,text='Dzuhur',font=changefont2).grid()

frame5 = LabelFrame(root,padx=5,pady=1)
frame5.place(x=80,y=460)
label5 = Label(frame5,text='Ashar',font=changefont2).grid()

frame6 = LabelFrame(root,padx=5,pady=1)
frame6.place(x=225,y=460)
label6 = Label(frame6,text='Maghrib',font=changefont2).grid()

frame7 = LabelFrame(root,padx=10,pady=1)
frame7.place(x=382,y=460)
label7 = Label(frame7,text='Isya',font=changefont2).grid()

# Tempat output 1 hari
frame8 = Entry(root,width=10,borderwidth=2) # Subuh
frame8.place(x=75,y=408)
frame9 = Entry(root,width=10,borderwidth=2) # Fajr
frame9.place(x=225,y=408)
frame10 = Entry(root,width=10,borderwidth=2) # Dzuhur
frame10.place(x=375,y=408)
frame11 = Entry(root,width=10,borderwidth=2) # Ashar
frame11.place(x=75,y=488)
frame12 = Entry(root,width=10,borderwidth=2) # Maghrib
frame12.place(x=225,y=488)
frame13 = Entry(root,width=10,borderwidth=2) # Isya
frame13.place(x=375,y=488)
frame15 = Entry(root,width=10,borderwidth=2)
frame15.place(x=60,y=335)
frame14 = Entry(root,width=10,borderwidth=2)
frame14.place(x=365,y=335)

# Opsi Pilihan tgl, bulan, tahun, kota
optionstahun_masehi = [1970,1971,1972,1973,1974,1975,1976,1977,1978,1979,1980,
	1981,1982,1983,1984,1985,1986,1987,1988,1989,1990,1991,1992,1993,1994,1995,
	1996,1997,1998,1999,2000,2001,2002,2003,2004,2005,2006,2007,2008,2009,2010,
	2011,2012,2013,2014,2015,2016,2017,2018,2019,2020,2021,2022,2023,2024,2025,
	2026,2027,2028,2029,2030,2031,2032,2033,2034,2035,2036,2037,2038,2039,2040]

optionsbulan_masehi = ['Januari','Februari','Maret','April','Mei','Juni', 
	'Juli','Agustus','September','Oktober','November','Desember']

optionstahun_hijriah = [1390,1391,1392,1393,1394,1395,1396,1397,1398,1399,1400,
	1401,1402,1403,1404,1405,1406,1407,1408,1409,1410,1411,1412,1413,1414,1415,
	1416,1417,1418,1419,1420,1421,1422,1423,1424,1425,1426,1427,1428,1429,1430,
	1431,1432,1433,1434,1435,1436,1437,1438,1439,1440,1441,1442,1443,1444,1445,
	1446,1447,1448,1449,1450,1451,1452,1453,1454,1455,1456,1457,1458,1459,1460]

optionsbulan_hijriah = ['Muharram','Safar','Rabiul Awwal','Rabiul Akhir','Jumadil Awwal',
	'Jumadil Akhir','Rajab','Syaban','Ramadhan','Syawal','Dzulqaidah','Dzulhijjah']

options1 = [1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31]
options2 = [1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30]
options3 = [1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29]
options4 = [1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28]

optionskota = ['Ambon', 'Banda Aceh', 'Bandar Lampung', 'Bandung', 'Banjarmasin', 'Bengkulu', 'Bima', 'Bogor', 'Denpasar',
	'Depok', 'Dompu', 'Gorontalo', 'Jakarta', 'Jambi', 'Jayapura', 'Kendari', 'Kupang', 'Makassar', 'Malang',
	'Manado', 'Manokwari', 'Mamuju', 'Mataram', 'Medan', 'Merauke', 'Padang', 'Palangkaraya', 'Palembang',
	'Pekanbaru', 'Palu', 'Pangkal Pinang', 'Pontianak', 'Praya', 'Sabang', 'Samarinda', 'Semarang', 'Serang',
	'Serui', 'Sorong', 'Sumbawa Besar', 'Surabaya', 'Tanjung Selor', 'Tanjung Pinang', 'Ternate', 'Yogyakarta']

def clicked(value):
	global tahun
	if value == 1:
		tahun_masehi.config(state='enable')
		tahun_hijriah.config(state='disable')
		tahun = tahun_masehi
	elif value == 2:
		tahun_masehi.config(state='disable')
		tahun_hijriah.config(state='enable')
		tahun = tahun_hijriah

def month(object):
	global bulan
	if tahun == tahun_masehi:
		if int(tahun_masehi.get()) % 4 == 0 and int(tahun_masehi.get()) % 400 == 0:
			mth_masehi = Label(root,text='Bulan: ')
			mth_masehi.place(x=135,y=45)
			bulan_masehi = ttk.Combobox(root,value=optionsbulan_masehi,width=15)
			bulan_masehi.bind('<<ComboboxSelected>>',kabisat_masehi)
			bulan_masehi.place(x=135,y=65)
			bulan = bulan_masehi
		else:
			mth_masehi = Label(root,text='Bulan: ')
			mth_masehi.place(x=135,y=45)
			bulan_masehi = ttk.Combobox(root,value=optionsbulan_masehi,width=15)
			bulan_masehi.bind('<<ComboboxSelected>>',nonkabisat_masehi)
			bulan_masehi.place(x=135,y=65)
			bulan = bulan_masehi
	elif tahun == tahun_hijriah:
		if int(tahun_hijriah.get()) % 30 == 2 or int(tahun_hijriah.get()) % 30 == 5 or int(tahun_hijriah.get()) % 30 == 7 or int(tahun_hijriah.get()) % 30 == 10 or int(tahun_hijriah.get()) % 30 == 13 or int(tahun_hijriah.get()) % 30 == 16 or int(tahun_hijriah.get()) % 30 == 18 or int(tahun_hijriah.get()) % 30 == 21 or int(tahun_hijriah.get()) % 30 == 24 or int(tahun_hijriah.get()) % 30 == 26 or int(tahun_hijriah.get()) % 30 == 29:
			mth_hijriah = Label(root,text='Bulan: ')
			mth_hijriah.place(x=135,y=120)
			bulan_hijriah = ttk.Combobox(root,value=optionsbulan_hijriah,width=15)
			bulan_hijriah.bind('<<ComboboxSelected>>',kabisat_hijriah)
			bulan_hijriah.place(x=135,y=140)
			bulan = bulan_hijriah
		else:
			mth_hijriah = Label(root,text='Bulan: ')
			mth_hijriah.place(x=135,y=120)
			bulan_hijriah = ttk.Combobox(root,value=optionsbulan_hijriah,width=15)
			bulan_hijriah.bind('<<ComboboxSelected>>',nonkabisat_hijriah)
			bulan_hijriah.place(x=135,y=140)
			bulan = bulan_hijriah

def kabisat_masehi(object):
	global tanggal
	if bulan.get() == 'Januari' or bulan.get() == 'Maret' or bulan.get() == 'Mei' or bulan.get() == 'Juli' or bulan.get() == 'Agustus' or bulan.get() == 'Oktober' or bulan.get() == 'Desember':
		date_masehi = Label(root,text='Tanggal: ')
		date_masehi.place(x=255,y=45)
		tanggal_masehi = ttk.Combobox(root,value=options1,width=10)
		tanggal_masehi.bind('<<ComboboxSelected>>')
		tanggal_masehi.place(x=255,y=65)
		tanggal = tanggal_masehi
	elif bulan.get() == 'Februari':
		date_masehi = Label(root,text='Tanggal: ')
		date_masehi.place(x=255,y=45)
		tanggal_masehi = ttk.Combobox(root,value=options3,width=10)
		tanggal_masehi.bind('<<ComboboxSelected>>')
		tanggal_masehi.place(x=255,y=65)
		tanggal = tanggal_masehi
	else:
		date_masehi = Label(root,text='Tanggal: ')
		date_masehi.place(x=255,y=45)
		tanggal_masehi = ttk.Combobox(root,value=options2,width=10)
		tanggal_masehi.bind('<<ComboboxSelected>>')
		tanggal_masehi.place(x=255,y=65)
		tanggal = tanggal_masehi

def nonkabisat_masehi(object):
	global tanggal
	if bulan.get() == 'Januari' or bulan.get() == 'Maret' or bulan.get() == 'Mei' or bulan.get() == 'Juli' or bulan.get() == 'Agustus' or bulan.get() == 'Oktober' or bulan.get() == 'Desember':
		date_masehi = Label(root,text='Tanggal: ')
		date_masehi.place(x=255,y=45)
		tanggal_masehi = ttk.Combobox(root,value=options1,width=10)
		tanggal_masehi.bind('<<ComboboxSelected>>')
		tanggal_masehi.place(x=255,y=65)
		tanggal = tanggal_masehi
	elif bulan.get() == 'Februari':
		date_masehi = Label(root,text='Tanggal: ')
		date_masehi.place(x=255, y=45)
		tanggal_masehi = ttk.Combobox(root,value=options4,width=10)
		tanggal_masehi.bind('<<ComboboxSelected>>')
		tanggal_masehi.place(x=255,y=65)
		tanggal = tanggal_masehi
	else:
		date_masehi = Label(root,text='Tanggal: ')
		date_masehi.place(x=255,y=45)
		tanggal_masehi = ttk.Combobox(root,value=options2,width=10)
		tanggal_masehi.bind('<<ComboboxSelected>>')
		tanggal_masehi.place(x=255,y=65)
		tanggal = tanggal_masehi

def kabisat_hijriah(object):
	global tanggal
	if bulan.get() == 'Muharram' or bulan.get() == 'Rabiul Awwal' or bulan.get() == 'Jumadil Awwal' or bulan.get() == 'Rajab' or bulan.get() == 'Dzulqaidah' or bulan.get() == 'Dzulhijjah':
		date_hijriah = Label(root,text='Tanggal: ')
		date_hijriah.place(x=255,y=120)
		tanggal_hijriah = ttk.Combobox(root,value=options2,width=10)
		tanggal_hijriah.bind('<<ComboboxSelected>>')
		tanggal_hijriah.place(x=255,y=140)
		tanggal = tanggal_hijriah
	else:
		date_hijriah = Label(root,text='Tanggal: ')
		date_hijriah.place(x=255,y=120)
		tanggal_hijriah = ttk.Combobox(root,value=options3,width=10)
		tanggal_hijriah.bind('<<ComboboxSelected>>')
		tanggal_hijriah.place(x=255,y=140)
		tanggal = tanggal_hijriah
		
def nonkabisat_hijriah(object):
	global tanggal
	if bulan.get() == 'Muharram' or bulan.get() == 'Rabiul Awwal' or bulan.get() == 'Jumadil Awwal' or bulan.get() == 'Rajab' or bulan.get() == 'Dzulqaidah':
		date_hijriah = Label(root,text='Tanggal: ')
		date_hijriah.place(x=255,y=120)
		tanggal_hijriah = ttk.Combobox(root,value=options2,width=10)
		tanggal_hijriah.bind('<<ComboboxSelected>>')
		tanggal_hijriah.place(x=255,y=140)
		tanggal = tanggal_hijriah
	else:
		date_hijriah = Label(root,text='Tanggal: ')
		date_hijriah.place(x=255,y=120)
		tanggal_hijriah = ttk.Combobox(root,value=options3,width=10)
		tanggal_hijriah.bind('<<ComboboxSelected>>')
		tanggal_hijriah.place(x=255,y=140)
		tanggal = tanggal_hijriah

yr_masehi = Label(root,text='Tahun: ')
yr_masehi.place(x=45,y=45)
tahun_masehi = ttk.Combobox(root,value=optionstahun_masehi,width=10,state='disable')
tahun_masehi.bind('<<ComboboxSelected>>',month)
tahun_masehi.set('2021')
tahun_masehi.place(x=45,y=65)

yr_hijriah = Label(root,text='Tahun: ')
yr_hijriah.place(x=45,y=120)
tahun_hijriah = ttk.Combobox(root,value=optionstahun_hijriah,width=10,state='disable')
tahun_hijriah.bind('<<ComboboxSelected>>',month)
tahun_hijriah.set('1443')
tahun_hijriah.place(x=45,y=140)

r = IntVar()
r.set('0')

masehi = Radiobutton(root,
	text='Masehi',variable=r,value=1,command=lambda:clicked(r.get())).place(x=25,y=20)
hijriah = Radiobutton(root,
	text='Hijriah',variable=r,value=2,command=lambda:clicked(r.get())).place(x=25,y=95)

kt = Label(root,text='Kota: ')
kt.place(x=30,y=180)
kota = ttk.Combobox(root,value=optionskota)
kota.bind('<<ComboboxSelected>>')
kota.place(x=30,y=200)

def cetak():
	class waktu_shalat:
		def __init__(self,tanggal,bulan,tahun,kota):
			self.D = tanggal
			self.M = bulan
			self.Y = tahun
			self.K = kota

		def hasil(self):
			data = load_workbook(filename="Test.xlsx")
			data.sheetnames
			sheet = data.active
			D = int(self.D)
			M = self.M
			Y = int(self.Y)
			K = self.K

			# Perhitungan 1 Hari
			# Perhitungan JD Masehi
			if M == 'Januari' or M == 'Februari' or M == 'Maret' or M == 'April' or M == 'Mei' or M == 'Juni' or M == 'Juli' or M == 'Agustus' or M == 'September' or M == 'Oktober' or M == 'November' or M == 'Desember':
				if M == 'Januari':
					Bln = 13
					Th = Y-1
				elif M == 'Februari':
					Bln = 14
					Th = Y-1
				elif M == 'Maret':
					Bln = 3
					Th = Y
				elif M == 'April':
					Bln = 4
					Th = Y
				elif M == 'Mei':
					Bln = 5
					Th = Y
				elif M == 'Juni':
					Bln = 6
					Th = Y
				elif M == 'Juli':
					Bln = 7
					Th = Y
				elif M == 'Agustus':
					Bln = 8
					Th = Y
				elif M == 'September':
					Bln = 9
					Th = Y
				elif M == 'Oktober':
					Bln = 10
					Th = Y
				elif M == 'November':
					Bln = 11
					Th = Y
				elif M == 'Desember':
					Bln = 12
					Th = Y

				A = int(Th/100)
				B = 2 + int(A/4) - A
				if Y < 1582:
					B = 0
				
				JD = 1720994.5 + int(365.25 * Th) + int(30.6001 * (Bln + 1)) + B + D

			# Perhitungan JD Hijriah
			elif M == 'Muharram' or M == 'Safar' or M == 'Rabiul Awwal' or M == 'Rabiul Akhir' or M == 'Jumadil Awwal' or M == 'Jumadil Akhir' or M == 'Rajab' or M == 'Syaban' or M == 'Ramadhan' or M == 'Syawal' or M == 'Dzulqaidah' or M == 'Dzulhijjah':
				if M == 'Muharram':
					M1 = 1
				elif M == 'Safar':
					M1 = 2
				elif M == 'Rabiul Awwal':
					M1 = 3
				elif M == 'Rabiul Akhir':
					M1 = 4
				elif M == 'Jumadil Awwal':
					M1 = 5
				elif M == 'Jumadil Akhir':
					M1 = 6
				elif M == 'Rajab':
					M1 = 7
				elif M == 'Syaban':
					M1 = 8
				elif M == 'Ramadhan':
					M1 = 9
				elif M == 'Syawal':
					M1 = 10
				elif M == 'Dzulqaidah':
					M1 = 11
				elif M == 'Dzulhijjah':
					M1 = 12

				Y1 = Y-1
				Y2 = int(Y1/30)
				Y3 = Y1-(Y2*30)
				if Y3 < 2:
				  k = 0
				elif Y3 < 5:
				  k = 1
				elif Y3 < 7:
				  k = 2			  
				elif Y3 < 10:
				  k = 3
				elif Y3 < 13:
				  k = 4
				elif Y3 < 16:
				  k = 5
				elif Y3 < 19:
				  k = 6
				elif Y3 < 21:
				  k = 7
				elif Y3 < 24:
				  k = 8
				elif Y3 < 26:
				  k = 9
				elif Y3 < 29:
				  k = 10
				else:
				  k = 11

				#Menghitung jumlah hari dalam tahun
				D1 = Y2*10631
				D2 = Y3*354+k

				#Mencari jumlah hari pada bulan genap dan ganjil
				M2 = M1-1
				if M1 %2 == 0:
				  D3 = ((M2/2)*30)+((M2/2)*29)
				else:
				  D3 = (((M2+1)/2)*30)+(((M2-1)/2)*29)

				#Jumlah hari
				D4 = D+D1+D2+D3
				JD = 1948438.5+D4

			# Fungsi Keluaran untuk 1 Hari
			def hasilWIB(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1):
				lbl0 = Label(frame14,text=K,font=changefont2).grid()
				lbl1 = Label(frame15,text=str(D)+' '+M+' '+str(Y),font=changefont2).grid()
				lbl2 = Label(frame8,text=str(WSJ)+':'+str(WSM1)+' WIB',font=changefont1).grid()
				lbl3 = Label(frame9,text=str(WTJ)+':'+str(WTM1)+' WIB',font=changefont1).grid()
				lbl4 = Label(frame10,text=str(WZJ)+':'+str(WZM2)+' WIB',font=changefont1).grid()
				lbl5 = Label(frame11,text=str(WAJ)+':'+str(WAM1)+' WIB',font=changefont1).grid()
				lbl6 = Label(frame12,text=str(WMJ)+':'+str(WMM1)+' WIB',font=changefont1).grid()
				lbl7 = Label(frame13,text=str(WIJ)+':'+str(WIM1)+' WIB',font=changefont1).grid()
				return(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)

			def hasilWITA(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1):
				lbl0 = Label(frame14,text=K,font=changefont2).grid()
				lbl1 = Label(frame15,text=str(D)+' '+M+' '+str(Y),font=changefont2).grid()
				lbl2 = Label(frame8,text=str(WSJ)+':'+str(WSM1)+' WITA',font=changefont1).grid()
				lbl3 = Label(frame9,text=str(WTJ)+':'+str(WTM1)+' WITA',font=changefont1).grid()
				lbl4 = Label(frame10,text=str(WZJ)+':'+str(WZM2)+' WITA',font=changefont1).grid()
				lbl5 = Label(frame11,text=str(WAJ)+':'+str(WAM1)+' WITA',font=changefont1).grid()
				lbl6 = Label(frame12,text=str(WMJ)+':'+str(WMM1)+' WITA',font=changefont1).grid()
				lbl7 = Label(frame13,text=str(WIJ)+':'+str(WIM1)+' WITA',font=changefont1).grid()
				return(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)

			def hasilWIT(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1):
				lbl0 = Label(frame14,text=K,font=changefont2).grid()
				lbl1 = Label(frame15,text=str(D)+' '+M+' '+str(Y),font=changefont2).grid()
				lbl2 = Label(frame8,text=str(WSJ)+':'+str(WSM1)+' WIT',font=changefont1).grid()
				lbl3 = Label(frame9,text=str(WTJ)+':'+str(WTM1)+' WIT',font=changefont1).grid()
				lbl4 = Label(frame10,text=str(WZJ)+':'+str(WZM2)+' WIT',font=changefont1).grid()
				lbl5 = Label(frame11,text=str(WAJ)+':'+str(WAM1)+' WIT',font=changefont1).grid()
				lbl6 = Label(frame12,text=str(WMJ)+':'+str(WMM1)+' WIT',font=changefont1).grid()
				lbl7 = Label(frame13,text=str(WIJ)+':'+str(WIM1)+' WIT',font=changefont1).grid()
				return(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)
			
			if K == sheet['A2'].value:
				L = sheet['B2'].value
				Bjr = sheet['C2'].value
				Z = sheet['D2'].value
				H = sheet['E2'].value

			elif K == sheet['A3'].value:
				L = sheet['B3'].value
				Bjr = sheet['C3'].value
				Z = sheet['D3'].value
				H = sheet['E3'].value

			elif K == sheet['A4'].value:
				L = sheet['B4'].value
				Bjr = sheet['C4'].value
				Z = sheet['D4'].value
				H = sheet['E4'].value

			elif K == sheet['A5'].value:
				L = sheet['B5'].value
				Bjr = sheet['C5'].value
				Z = sheet['D5'].value
				H = sheet['E5'].value

			elif K == sheet['A6'].value:
				L = sheet['B6'].value
				Bjr = sheet['C6'].value
				Z = sheet['D6'].value
				H = sheet['E6'].value

			elif K == sheet['A7'].value:
				L = sheet['B7'].value
				Bjr = sheet['C7'].value
				Z = sheet['D7'].value
				H = sheet['E7'].value

			elif K == sheet['A8'].value:
				L = sheet['B8'].value
				Bjr = sheet['C8'].value
				Z = sheet['D8'].value
				H = sheet['E8'].value

			elif K == sheet['A9'].value:
				L = sheet['B9'].value
				Bjr = sheet['C9'].value
				Z = sheet['D9'].value
				H = sheet['E9'].value

			elif K == sheet['A10'].value:
				L = sheet['B10'].value
				Bjr = sheet['C10'].value
				Z = sheet['D10'].value
				H = sheet['E10'].value

			elif K == sheet['A11'].value:
				L = sheet['B11'].value
				Bjr = sheet['C11'].value
				Z = sheet['D11'].value
				H = sheet['E11'].value

			elif K == sheet['A12'].value:
				L = sheet['B12'].value
				Bjr = sheet['C12'].value
				Z = sheet['D12'].value
				H = sheet['E12'].value

			elif K == sheet['A13'].value:
				L = sheet['B13'].value
				Bjr = sheet['C13'].value
				Z = sheet['D13'].value
				H = sheet['E13'].value

			elif K == sheet['A14'].value:
				L = sheet['B14'].value
				Bjr = sheet['C14'].value
				Z = sheet['D14'].value
				H = sheet['E14'].value

			elif K == sheet['A15'].value:
				L = sheet['B15'].value
				Bjr = sheet['C15'].value
				Z = sheet['D15'].value
				H = sheet['E15'].value

			elif K == sheet['A16'].value:
				L = sheet['B16'].value
				Bjr = sheet['C16'].value
				Z = sheet['D16'].value
				H = sheet['E16'].value

			elif K == sheet['A17'].value:
				L = sheet['B17'].value
				Bjr = sheet['C17'].value
				Z = sheet['D17'].value
				H = sheet['E17'].value

			elif K == sheet['A18'].value:
				L = sheet['B18'].value
				Bjr = sheet['C18'].value
				Z = sheet['D18'].value
				H = sheet['E18'].value

			elif K == sheet['A19'].value:
				L = sheet['B19'].value
				Bjr = sheet['C19'].value
				Z = sheet['D19'].value
				H = sheet['E19'].value

			elif K == sheet['A20'].value:
				L = sheet['B20'].value
				Bjr = sheet['C20'].value
				Z = sheet['D20'].value
				H = sheet['E20'].value

			elif K == sheet['A21'].value:
				L = sheet['B21'].value
				Bjr = sheet['C21'].value
				Z = sheet['D21'].value
				H = sheet['E21'].value

			elif K == sheet['A22'].value:
				L = sheet['B22'].value
				Bjr = sheet['C22'].value
				Z = sheet['D22'].value
				H = sheet['E22'].value

			elif K == sheet['A23'].value:
				L = sheet['B23'].value
				Bjr = sheet['C23'].value
				Z = sheet['D23'].value
				H = sheet['E23'].value

			elif K == sheet['A24'].value:
				L = sheet['B24'].value
				Bjr = sheet['C24'].value
				Z = sheet['D24'].value
				H = sheet['E24'].value

			elif K == sheet['A25'].value:
				L = sheet['B25'].value
				Bjr = sheet['C25'].value
				Z = sheet['D25'].value
				H = sheet['E25'].value

			elif K == sheet['A26'].value:
				L = sheet['B26'].value
				Bjr = sheet['C26'].value
				Z = sheet['D26'].value
				H = sheet['E26'].value

			elif K == sheet['A27'].value:
				L = sheet['B27'].value
				Bjr = sheet['C27'].value
				Z = sheet['D27'].value
				H = sheet['E27'].value

			elif K == sheet['A28'].value:
				L = sheet['B28'].value
				Bjr = sheet['C28'].value
				Z = sheet['D28'].value
				H = sheet['E28'].value

			elif K == sheet['A29'].value:
				L = sheet['B29'].value
				Bjr = sheet['C29'].value
				Z = sheet['D29'].value
				H = sheet['E29'].value

			elif K == sheet['A30'].value:
				L = sheet['B30'].value
				Bjr = sheet['C30'].value
				Z = sheet['D30'].value
				H = sheet['E30'].value

			elif K == sheet['A31'].value:
				L = sheet['B31'].value
				Bjr = sheet['C31'].value
				Z = sheet['D31'].value
				H = sheet['E31'].value

			elif K == sheet['A32'].value:
				L = sheet['B32'].value
				Bjr = sheet['C32'].value
				Z = sheet['D32'].value
				H = sheet['E32'].value

			elif K == sheet['A33'].value:
				L = sheet['B33'].value
				Bjr = sheet['C33'].value
				Z = sheet['D33'].value
				H = sheet['E33'].value

			elif K == sheet['A34'].value:
				L = sheet['B34'].value
				Bjr = sheet['C34'].value
				Z = sheet['D34'].value
				H = sheet['E34'].value

			elif K == sheet['A35'].value:
				L = sheet['B35'].value
				Bjr = sheet['C35'].value
				Z = sheet['D35'].value
				H = sheet['E35'].value

			elif K == sheet['A36'].value:
				L = sheet['B36'].value
				Bjr = sheet['C36'].value
				Z = sheet['D36'].value
				H = sheet['E36'].value

			elif K == sheet['A37'].value:
				L = sheet['B37'].value
				Bjr = sheet['C37'].value
				Z = sheet['D37'].value
				H = sheet['E37'].value

			elif K == sheet['A38'].value:
				L = sheet['B38'].value
				Bjr = sheet['C38'].value
				Z = sheet['D38'].value
				H = sheet['E38'].value

			elif K == sheet['A39'].value:
				L = sheet['B39'].value
				Bjr = sheet['C39'].value
				Z = sheet['D39'].value
				H = sheet['E39'].value

			elif K == sheet['A40'].value:
				L = sheet['B40'].value
				Bjr = sheet['C40'].value
				Z = sheet['D40'].value
				H = sheet['E40'].value

			elif K == sheet['A41'].value:
				L = sheet['B41'].value
				Bjr = sheet['C41'].value
				Z = sheet['D41'].value
				H = sheet['E41'].value

			elif K == sheet['A42'].value:
				L = sheet['B42'].value
				Bjr = sheet['C42'].value
				Z = sheet['D42'].value
				H = sheet['E42'].value

			elif K == sheet['A43'].value:
				L = sheet['B43'].value
				Bjr = sheet['C43'].value
				Z = sheet['D43'].value
				H = sheet['E43'].value

			elif K == sheet['A44'].value:
				L = sheet['B44'].value
				Bjr = sheet['C44'].value
				Z = sheet['D44'].value
				H = sheet['E44'].value

			elif K == sheet['A45'].value:
				L = sheet['B45'].value
				Bjr = sheet['C45'].value
				Z = sheet['D45'].value
				H = sheet['E45'].value

			elif K == sheet['A46'].value:
				L = sheet['B46'].value
				Bjr = sheet['C46'].value
				Z = sheet['D46'].value
				H = sheet['E46'].value

			AS = -20
			AI = -18
			KA = 1

			JD1 = JD - Z/24
			T = 2*3.14*(JD1 - 2451545)/365.25
			Delta = 0.3787 + 23.64*sin(radians(57.297*T - 79.547)) + 0.3812*sin(radians(2*57.297*T - 82.682)) + 0.17132*sin(radians(3*57.297*T - 59.722))
			U = (JD1 - 2451545)/36525
			L0 = (280.46607+36000.7698*U)
			# Membuat L0 Menjadi Sudut Antara 0 - 360
			if L0 < 360:
				L0 = L0
			elif L0 < 720:
				L0 = L0 - 360
			elif L0 < 1081:
				L0 = L0 - 720
			elif L0 < 1441:
				L0 = L0 - 1080
			elif L0 < 1801:
				L0 = L0 - 1440
			elif L0 < 2161:
				L0 = L0 - 1800
			elif L0 < 2521:
				L0 = L0 - 2160
			elif L0 < 2881:
				L0 = L0 - 2520
			elif L0 < 3241:
				L0 = L0 - 2880
			elif L0 < 3601:
				L0 = L0 - 3240
			elif L0 < 3960:
				L0 = L0 - 3600

			# Waktu Transit
			ET = (-1*(1789+237*U)*sin(radians(L0))-(7416-62*U)*cos(radians(L0))+(9934-14*U)*sin(radians(2*L0))-(29+5*U)*cos(radians(2*L0))+(74+10*U)*sin(radians(3*L0))+(320-4*U)*cos(radians(3*L0))-212*sin(radians(4*L0)))/1000
			Transit = 12 + Z - (Bjr/15) - (ET/60)

			# Waktu Zuhur
			WZJ = int(Transit)
			WZ2 = Transit - WZJ
			WZM = (int(WZ2*60))
			WZ3 = ((WZ2*60) - WZM)
			WZS = int(WZ3*60)
			WZM1 = WZM + 2
			if WZS > 0:
				WZM2 = WZM1 + 1
			else:
				WZM2 = WZM1

			if WZM2 >= 60:
				WZJ = WZJ + 1
				WZM2 = WZM2 - 60
			else:
				WZJ = WZJ
				WZM2 = WZM2

			#Waktu Ashar
			AA = acot(KA + tan(radians(abs(Delta - L))))
			AA1 = degrees(AA)
			HAA = acos((sin(radians(AA1)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
			HAA1 = degrees(HAA)
			WA = Transit + (HAA1/15)
			WAJ = int(WA)
			WA2 = WA - WAJ
			WAM = int(WA2*60)
			WA3 = ((WA2*60) - WAM)
			WAS = int(WA3*60)
			if WAS > 0:
				WAM1 = WAM + 1
			else:
				WAM1 = WAM

			if WAM1 >= 60:
				WAJ = WAJ + 1
				WAM1 = WAM1 - 60
			else:
				WAJ = WAJ
				WAM1 = WAM1

			# Waktu Maghrib
			AM = (-0.8333) - 0.0347*sqrt(H)
			HAM = acos((sin(radians(AM)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
			HAM1 = degrees(HAM)
			WM = Transit + (HAM1/15)
			WMJ = int(WM)
			WM2 = WM - WMJ
			WMM = int(WM2*60)
			WM3 = ((WM2*60) - WMM)
			WMS = int(WM3*60)
			if WMS > 0:
				WMM1 = WMM + 1
			else:
				WMM1 = WMM

			if WMM1 >= 60:
				WMJ = WMJ + 1
				WMM1 = WMM1 - 60
			else:
				WMJ = WMJ
				WMM1 = WMM1

			# Waktu Isya'
			HAI = acos((sin(radians(AI)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
			HAI1 = degrees(HAI)
			WI = Transit + (HAI1/15)
			WIJ = int(WI)
			WI2 = WI - WIJ
			WIM = int(WI2*60)
			WI3 = ((WI2*60) - WIM)
			WIS = int(WI3*60)
			if WIS > 0:
				WIM1 = WIM + 1
			else:
				WIM1 = WIM

			if WIM1 >= 60:
				WIJ = WIJ + 1
				WIM1 = WIM1 - 60
			else:
				WIJ = WIJ
				WIM1 = WIM1

			# Waktu Subuh
			HAS = acos((sin(radians(AS)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
			HAS1 = degrees(HAS)
			WS = Transit - (HAS1/15)
			WSJ = int(WS)
			WS2 = WS - WSJ
			WSM = int(WS2*60)
			WS3 = ((WS2*60) - WSM)
			WSS = int(WS3*60)
			if WSS > 0:
				WSM1 = WSM + 1
			else:
				WSM1 = WSM

			if WSM1 >= 60:
				WSJ = WSJ + 1
				WSM1 = WSM1 - 60
			else:
				WSJ = WSJ
				WSM1 = WSM1

			# Waktu Terbit Matahari
			AT = (-0.8333) - 0.0347*sqrt(H)
			HAT = acos((sin(radians(AT)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
			HAT1 = degrees(HAT)
			WT = Transit - (HAT1/15)
			WTJ = int(WT)
			WT2 = WT - WTJ
			WTM = int(WT2*60)
			WT3 = ((WT2*60) - WTM)
			WTS = int(WT3*60)
			if WTS > 0:
				WTM1 = WTM + 1
			else:
				WTM1 = WTM

			if WTM1 >= 60:
				WTJ = WTJ + 1
				WTM1 = WTM1 - 60
			else:
				WTJ = WTJ
				WTM1 = WTM1

			if Z == 7:
				(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1) = hasilWIB(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)
			elif Z == 8:
				(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1) = hasilWITA(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)
			elif Z ==9:
				(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1) = hasilWIT(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)

			# Fungsi Keluaran Untuk 1 Bulan
			def satubulanWIB(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1):
				count = D-1
				ktout = Label(root,text=K,font=changefont)
				ktout.place(x=975,y=15)
				wout = Label(root,text=str(M)+' '+str(Y),font=changefont)
				wout.place(x=560,y=15)
				lblsubuh = str(WSJ),':',str(WSM1)
				lblfajr = str(WTJ),':',str(WTM1)
				lblzuhur = str(WZJ),':',str(WZM2)
				lblashar = str(WAJ),':',str(WAM1)
				lblmaghrib = str(WMJ),':',str(WMM1)
				lblisya = str(WIJ),':',str(WIM1)
				tabel.insert(parent='',index='end',iid=count,text='',values=(D,lblsubuh,lblfajr,lblzuhur,lblashar,lblmaghrib,lblisya,'WIB'))
				return(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)

			def satubulanWITA(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1):
				count = D-1
				ktout = Label(root,text=K,font=changefont)
				ktout.place(x=975,y=15)
				wout = Label(root,text=str(M)+' '+str(Y),font=changefont)
				wout.place(x=560,y=15)
				lblsubuh = str(WSJ),':',str(WSM1)
				lblfajr = str(WTJ),':',str(WTM1)
				lblzuhur = str(WZJ),':',str(WZM2)
				lblashar = str(WAJ),':',str(WAM1)
				lblmaghrib = str(WMJ),':',str(WMM1)
				lblisya = str(WIJ),':',str(WIM1)
				tabel.insert(parent='',index='end',iid=count,text='',values=(D,lblsubuh,lblfajr,lblzuhur,lblashar,lblmaghrib,lblisya,'WITA'))
				return(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)

			def satubulanWIT(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1):
				count = D-1
				ktout = Label(root,text=K,font=changefont)
				ktout.place(x=975,y=15)
				wout = Label(root,text=str(M)+' '+str(Y),font=changefont)
				wout.place(x=560,y=15)
				lblsubuh = str(WSJ),':',str(WSM1)
				lblfajr = str(WTJ),':',str(WTM1)
				lblzuhur = str(WZJ),':',str(WZM2)
				lblashar = str(WAJ),':',str(WAM1)
				lblmaghrib = str(WMJ),':',str(WMM1)
				lblisya = str(WIJ),':',str(WIM1)
				tabel.insert(parent='',index='end',iid=count,text='',values=(D,lblsubuh,lblfajr,lblzuhur,lblashar,lblmaghrib,lblisya,'WIT'))
				return(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)

			# Perhitungan 1 Bulan Masehi
			if M == 'Januari' or M == 'Februari' or M == 'Maret' or M == 'April' or M == 'Mei' or M == 'Juni' or M == 'Juli' or M == 'Agustus' or M == 'September' or M == 'Oktober' or M == 'November' or M == 'Desember':
				# 1 Bulan Kabisat
				if Y % 4 == 0 and Y % 400 == 0:
					if Bln == 13 or Bln == 3 or Bln == 5 or Bln == 7 or Bln == 8 or Bln == 10 or Bln == 12:
						for D in range(1,32):
							JD = 1720994.5 + int(365.25 * Th) + int(30.6001 * (Bln + 1)) + B + D
							JD1 = JD - Z/24
							T = 2*3.14*(JD1 - 2451545)/365.25
							Delta = 0.3787 + 23.64*sin(radians(57.297*T - 79.547)) + 0.3812*sin(radians(2*57.297*T - 82.682)) + 0.17132*sin(radians(3*57.297*T - 59.722))
							U = (JD1 - 2451545)/36525
							L0 = (280.46607+36000.7698*U)
							if L0 < 360:
								L0 = L0
							elif L0 < 720:
								L0 = L0 - 360
							elif L0 < 1081:
								L0 = L0 - 720
							elif L0 < 1441:
								L0 = L0 - 1080
							elif L0 < 1801:
								L0 = L0 - 1440
							elif L0 < 2161:
								L0 = L0 - 1800
							elif L0 < 2521:
								L0 = L0 - 2160
							elif L0 < 2881:
								L0 = L0 - 2520
							elif L0 < 3241:
								L0 = L0 - 2880
							elif L0 < 3601:
								L0 = L0 - 3240
							elif L0 < 3960:
								L0 = L0 - 3600

							# Waktu Transit
							ET = (-1*(1789+237*U)*sin(radians(L0))-(7416-62*U)*cos(radians(L0))+(9934-14*U)*sin(radians(2*L0))-(29+5*U)*cos(radians(2*L0))+(74+10*U)*sin(radians(3*L0))+(320-4*U)*cos(radians(3*L0))-212*sin(radians(4*L0)))/1000
							Transit = 12 + Z - (Bjr/15) - (ET/60)

							# Waktu Zuhur
							WZJ = int(Transit)
							WZ2 = Transit - WZJ
							WZM = (int(WZ2*60))
							WZ3 = ((WZ2*60) - WZM)
							WZS = int(WZ3*60)
							WZM1 = WZM + 2
							if WZS > 0:
								WZM2 = WZM1 + 1
							else:
								WZM2 = WZM1

							if WZM2 >= 60:
								WZJ = WZJ + 1
								WZM2 = WZM2 - 60
							else:
								WZJ = WZJ
								WZM2 = WZM2

							#Waktu Ashar
							AA = acot(KA + tan(radians(abs(Delta - L))))
							AA1 = degrees(AA)
							HAA = acos((sin(radians(AA1)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAA1 = degrees(HAA)
							WA = Transit + (HAA1/15)
							WAJ = int(WA)
							WA2 = WA - WAJ
							WAM = int(WA2*60)
							WA3 = ((WA2*60) - WAM)
							WAS = int(WA3*60)
							if WAS > 0:
								WAM1 = WAM + 1
							else:
								WAM1 = WAM

							if WAM1 >= 60:
								WAJ = WAJ + 1
								WAM1 = WAM1 - 60
							else:
								WAJ = WAJ
								WAM1 = WAM1

							# Waktu Maghrib
							AM = (-0.8333) - 0.0347*sqrt(H)
							HAM = acos((sin(radians(AM)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAM1 = degrees(HAM)
							WM = Transit + (HAM1/15)
							WMJ = int(WM)
							WM2 = WM - WMJ
							WMM = int(WM2*60)
							WM3 = ((WM2*60) - WMM)
							WMS = int(WM3*60)
							if WMS > 0:
								WMM1 = WMM + 1
							else:
								WMM1 = WMM

							if WMM1 >= 60:
								WMJ = WMJ + 1
								WMM1 = WMM1 - 60
							else:
								WMJ = WMJ
								WMM1 = WMM1

							# Waktu Isya'
							HAI = acos((sin(radians(AI)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAI1 = degrees(HAI)
							WI = Transit + (HAI1/15)
							WIJ = int(WI)
							WI2 = WI - WIJ
							WIM = int(WI2*60)
							WI3 = ((WI2*60) - WIM)
							WIS = int(WI3*60)
							if WIS > 0:
								WIM1 = WIM + 1
							else:
								WIM1 = WIM

							if WIM1 >= 60:
								WIJ = WIJ + 1
								WIM1 = WIM1 - 60
							else:
								WIJ = WIJ
								WIM1 = WIM1

							# Waktu Subuh
							HAS = acos((sin(radians(AS)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAS1 = degrees(HAS)
							WS = Transit - (HAS1/15)
							WSJ = int(WS)
							WS2 = WS - WSJ
							WSM = int(WS2*60)
							WS3 = ((WS2*60) - WSM)
							WSS = int(WS3*60)
							if WSS > 0:
								WSM1 = WSM + 1
							else:
								WSM1 = WSM

							if WSM1 >= 60:
								WSJ = WSJ + 1
								WSM1 = WSM1 - 60
							else:
								WSJ = WSJ
								WSM1 = WSM1

							# Waktu Terbit Matahari
							AT = (-0.8333) - 0.0347*sqrt(H)
							HAT = acos((sin(radians(AT)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAT1 = degrees(HAT)
							WT = Transit - (HAT1/15)
							WTJ = int(WT)
							WT2 = WT - WTJ
							WTM = int(WT2*60)
							WT3 = ((WT2*60) - WTM)
							WTS = int(WT3*60)
							if WTS > 0:
								WTM1 = WTM + 1
							else:
								WTM1 = WTM

							if WTM1 >= 60:
								WTJ = WTJ + 1
								WTM1 = WTM1 - 60
							else:
								WTJ = WTJ
								WTM1 = WTM1

							if Z == 7:
								(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1) = satubulanWIB(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)
							elif Z == 8:
								(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1) = satubulanWITA(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)
							elif Z ==9:
								(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1) = satubulanWIT(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)
					elif Bln == 14:
						for D in range(1,30):
							JD = 1720994.5 + int(365.25 * Th) + int(30.6001 * (Bln + 1)) + B + D
							JD1 = JD - Z/24
							T = 2*3.14*(JD1 - 2451545)/365.25
							Delta = 0.3787 + 23.64*sin(radians(57.297*T - 79.547)) + 0.3812*sin(radians(2*57.297*T - 82.682)) + 0.17132*sin(radians(3*57.297*T - 59.722))
							U = (JD1 - 2451545)/36525
							L0 = (280.46607+36000.7698*U)

							# Membuat L0 Menjadi Sudut Antara 0 - 360
							if L0 < 360:
								L0 = L0
							elif L0 < 720:
								L0 = L0 - 360
							elif L0 < 1081:
								L0 = L0 - 720
							elif L0 < 1441:
								L0 = L0 - 1080
							elif L0 < 1801:
								L0 = L0 - 1440
							elif L0 < 2161:
								L0 = L0 - 1800
							elif L0 < 2521:
								L0 = L0 - 2160
							elif L0 < 2881:
								L0 = L0 - 2520
							elif L0 < 3241:
								L0 = L0 - 2880
							elif L0 < 3601:
								L0 = L0 - 3240
							elif L0 < 3960:
								L0 = L0 - 3600

							# Waktu Transit
							ET = (-1*(1789+237*U)*sin(radians(L0))-(7416-62*U)*cos(radians(L0))+(9934-14*U)*sin(radians(2*L0))-(29+5*U)*cos(radians(2*L0))+(74+10*U)*sin(radians(3*L0))+(320-4*U)*cos(radians(3*L0))-212*sin(radians(4*L0)))/1000
							Transit = 12 + Z - (Bjr/15) - (ET/60)

							# Waktu Zuhur
							WZJ = int(Transit)
							WZ2 = Transit - WZJ
							WZM = (int(WZ2*60))
							WZ3 = ((WZ2*60) - WZM)
							WZS = int(WZ3*60)
							WZM1 = WZM + 2
							if WZS > 0:
								WZM2 = WZM1 + 1
							else:
								WZM2 = WZM1

							if WZM2 >= 60:
								WZJ = WZJ + 1
								WZM2 = WZM2 - 60
							else:
								WZJ = WZJ
								WZM2 = WZM2

							#Waktu Ashar
							AA = acot(KA + tan(radians(abs(Delta - L))))
							AA1 = degrees(AA)
							HAA = acos((sin(radians(AA1)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAA1 = degrees(HAA)
							WA = Transit + (HAA1/15)
							WAJ = int(WA)
							WA2 = WA - WAJ
							WAM = int(WA2*60)
							WA3 = ((WA2*60) - WAM)
							WAS = int(WA3*60)
							if WAS > 0:
								WAM1 = WAM + 1
							else:
								WAM1 = WAM

							if WAM1 >= 60:
								WAJ = WAJ + 1
								WAM1 = WAM1 - 60
							else:
								WAJ = WAJ
								WAM1 = WAM1

							# Waktu Maghrib
							AM = (-0.8333) - 0.0347*sqrt(H)
							HAM = acos((sin(radians(AM)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAM1 = degrees(HAM)
							WM = Transit + (HAM1/15)
							WMJ = int(WM)
							WM2 = WM - WMJ
							WMM = int(WM2*60)
							WM3 = ((WM2*60) - WMM)
							WMS = int(WM3*60)
							if WMS > 0:
								WMM1 = WMM + 1
							else:
								WMM1 = WMM

							if WMM1 >= 60:
								WMJ = WMJ + 1
								WMM1 = WMM1 - 60
							else:
								WMJ = WMJ
								WMM1 = WMM1

							# Waktu Isya'
							HAI = acos((sin(radians(AI)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAI1 = degrees(HAI)
							WI = Transit + (HAI1/15)
							WIJ = int(WI)
							WI2 = WI - WIJ
							WIM = int(WI2*60)
							WI3 = ((WI2*60) - WIM)
							WIS = int(WI3*60)
							if WIS > 0:
								WIM1 = WIM + 1
							else:
								WIM1 = WIM

							if WIM1 >= 60:
								WIJ = WIJ + 1
								WIM1 = WIM1 - 60
							else:
								WIJ = WIJ
								WIM1 = WIM1

							# Waktu Subuh
							HAS = acos((sin(radians(AS)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAS1 = degrees(HAS)
							WS = Transit - (HAS1/15)
							WSJ = int(WS)
							WS2 = WS - WSJ
							WSM = int(WS2*60)
							WS3 = ((WS2*60) - WSM)
							WSS = int(WS3*60)
							if WSS > 0:
								WSM1 = WSM + 1
							else:
								WSM1 = WSM

							if WSM1 >= 60:
								WSJ = WSJ + 1
								WSM1 = WSM1 - 60
							else:
								WSJ = WSJ
								WSM1 = WSM1

							# Waktu Terbit Matahari
							AT = (-0.8333) - 0.0347*sqrt(H)
							HAT = acos((sin(radians(AT)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAT1 = degrees(HAT)
							WT = Transit - (HAT1/15)
							WTJ = int(WT)
							WT2 = WT - WTJ
							WTM = int(WT2*60)
							WT3 = ((WT2*60) - WTM)
							WTS = int(WT3*60)
							if WTS > 0:
								WTM1 = WTM + 1
							else:
								WTM1 = WTM

							if WTM1 >= 60:
								WTJ = WTJ + 1
								WTM1 = WTM1 - 60
							else:
								WTJ = WTJ
								WTM1 = WTM1

							if Z == 7:
								(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1) = satubulanWIB(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)
							elif Z == 8:
								(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1) = satubulanWITA(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)
							elif Z ==9:
								(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1) = satubulanWIT(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)
					else:
						for D in range(1,31):
							JD = 1720994.5 + int(365.25 * Th) + int(30.6001 * (Bln + 1)) + B + D
							JD1 = JD - Z/24
							T = 2*3.14*(JD1 - 2451545)/365.25
							Delta = 0.3787 + 23.64*sin(radians(57.297*T - 79.547)) + 0.3812*sin(radians(2*57.297*T - 82.682)) + 0.17132*sin(radians(3*57.297*T - 59.722))
							U = (JD1 - 2451545)/36525
							L0 = (280.46607+36000.7698*U)

							# Membuat L0 Menjadi Sudut Antara 0 - 360
							if L0 < 360:
								L0 = L0
							elif L0 < 720:
								L0 = L0 - 360
							elif L0 < 1081:
								L0 = L0 - 720
							elif L0 < 1441:
								L0 = L0 - 1080
							elif L0 < 1801:
								L0 = L0 - 1440
							elif L0 < 2161:
								L0 = L0 - 1800
							elif L0 < 2521:
								L0 = L0 - 2160
							elif L0 < 2881:
								L0 = L0 - 2520
							elif L0 < 3241:
								L0 = L0 - 2880
							elif L0 < 3601:
								L0 = L0 - 3240
							elif L0 < 3960:
								L0 = L0 - 3600

							# Waktu Transit
							ET = (-1*(1789+237*U)*sin(radians(L0))-(7416-62*U)*cos(radians(L0))+(9934-14*U)*sin(radians(2*L0))-(29+5*U)*cos(radians(2*L0))+(74+10*U)*sin(radians(3*L0))+(320-4*U)*cos(radians(3*L0))-212*sin(radians(4*L0)))/1000
							Transit = 12 + Z - (Bjr/15) - (ET/60)

							# Waktu Zuhur
							WZJ = int(Transit)
							WZ2 = Transit - WZJ
							WZM = (int(WZ2*60))
							WZ3 = ((WZ2*60) - WZM)
							WZS = int(WZ3*60)
							WZM1 = WZM + 2
							if WZS > 0:
								WZM2 = WZM1 + 1
							else:
								WZM2 = WZM1

							if WZM2 >= 60:
								WZJ = WZJ + 1
								WZM2 = WZM2 - 60
							else:
								WZJ = WZJ
								WZM2 = WZM2

							#Waktu Ashar
							AA = acot(KA + tan(radians(abs(Delta - L))))
							AA1 = degrees(AA)
							HAA = acos((sin(radians(AA1)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAA1 = degrees(HAA)
							WA = Transit + (HAA1/15)
							WAJ = int(WA)
							WA2 = WA - WAJ
							WAM = int(WA2*60)
							WA3 = ((WA2*60) - WAM)
							WAS = int(WA3*60)
							if WAS > 0:
								WAM1 = WAM + 1
							else:
								WAM1 = WAM

							if WAM1 >= 60:
								WAJ = WAJ + 1
								WAM1 = WAM1 - 60
							else:
								WAJ = WAJ
								WAM1 = WAM1

							# Waktu Maghrib
							AM = (-0.8333) - 0.0347*sqrt(H)
							HAM = acos((sin(radians(AM)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAM1 = degrees(HAM)
							WM = Transit + (HAM1/15)
							WMJ = int(WM)
							WM2 = WM - WMJ
							WMM = int(WM2*60)
							WM3 = ((WM2*60) - WMM)
							WMS = int(WM3*60)
							if WMS > 0:
								WMM1 = WMM + 1
							else:
								WMM1 = WMM

							if WMM1 >= 60:
								WMJ = WMJ + 1
								WMM1 = WMM1 - 60
							else:
								WMJ = WMJ
								WMM1 = WMM1

							# Waktu Isya'
							HAI = acos((sin(radians(AI)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAI1 = degrees(HAI)
							WI = Transit + (HAI1/15)
							WIJ = int(WI)
							WI2 = WI - WIJ
							WIM = int(WI2*60)
							WI3 = ((WI2*60) - WIM)
							WIS = int(WI3*60)
							if WIS > 0:
								WIM1 = WIM + 1
							else:
								WIM1 = WIM

							if WIM1 >= 60:
								WIJ = WIJ + 1
								WIM1 = WIM1 - 60
							else:
								WIJ = WIJ
								WIM1 = WIM1

							# Waktu Subuh
							HAS = acos((sin(radians(AS)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAS1 = degrees(HAS)
							WS = Transit - (HAS1/15)
							WSJ = int(WS)
							WS2 = WS - WSJ
							WSM = int(WS2*60)
							WS3 = ((WS2*60) - WSM)
							WSS = int(WS3*60)
							if WSS > 0:
								WSM1 = WSM + 1
							else:
								WSM1 = WSM

							if WSM1 >= 60:
								WSJ = WSJ + 1
								WSM1 = WSM1 - 60
							else:
								WSJ = WSJ
								WSM1 = WSM1

							# Waktu Terbit Matahari
							AT = (-0.8333) - 0.0347*sqrt(H)
							HAT = acos((sin(radians(AT)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAT1 = degrees(HAT)
							WT = Transit - (HAT1/15)
							WTJ = int(WT)
							WT2 = WT - WTJ
							WTM = int(WT2*60)
							WT3 = ((WT2*60) - WTM)
							WTS = int(WT3*60)
							if WTS > 0:
								WTM1 = WTM + 1
							else:
								WTM1 = WTM

							if WTM1 >= 60:
								WTJ = WTJ + 1
								WTM1 = WTM1 - 60
							else:
								WTJ = WTJ
								WTM1 = WTM1

							if Z == 7:
								(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1) = satubulanWIB(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)
							elif Z == 8:
								(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1) = satubulanWITA(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)
							elif Z ==9:
								(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1) = satubulanWIT(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)
				else:
					if Bln == 13 or Bln == 3 or Bln == 5 or Bln == 7 or Bln == 8 or Bln == 10 or Bln == 12:
						for D in range(1,32):
							JD = 1720994.5 + int(365.25 * Th) + int(30.6001 * (Bln + 1)) + B + D
							JD1 = JD - Z/24
							T = 2*3.14*(JD1 - 2451545)/365.25
							Delta = 0.3787 + 23.64*sin(radians(57.297*T - 79.547)) + 0.3812*sin(radians(2*57.297*T - 82.682)) + 0.17132*sin(radians(3*57.297*T - 59.722))
							U = (JD1 - 2451545)/36525
							L0 = (280.46607+36000.7698*U)

							# Membuat L0 Menjadi Sudut Antara 0 - 360
							if L0 < 360:
								L0 = L0
							elif L0 < 720:
								L0 = L0 - 360
							elif L0 < 1081:
								L0 = L0 - 720
							elif L0 < 1441:
								L0 = L0 - 1080
							elif L0 < 1801:
								L0 = L0 - 1440
							elif L0 < 2161:
								L0 = L0 - 1800
							elif L0 < 2521:
								L0 = L0 - 2160
							elif L0 < 2881:
								L0 = L0 - 2520
							elif L0 < 3241:
								L0 = L0 - 2880
							elif L0 < 3601:
								L0 = L0 - 3240
							elif L0 < 3960:
								L0 = L0 - 3600

							# Waktu Transit
							ET = (-1*(1789+237*U)*sin(radians(L0))-(7416-62*U)*cos(radians(L0))+(9934-14*U)*sin(radians(2*L0))-(29+5*U)*cos(radians(2*L0))+(74+10*U)*sin(radians(3*L0))+(320-4*U)*cos(radians(3*L0))-212*sin(radians(4*L0)))/1000
							Transit = 12 + Z - (Bjr/15) - (ET/60)

							# Waktu Zuhur
							WZJ = int(Transit)
							WZ2 = Transit - WZJ
							WZM = (int(WZ2*60))
							WZ3 = ((WZ2*60) - WZM)
							WZS = int(WZ3*60)
							WZM1 = WZM + 2
							if WZS > 0:
								WZM2 = WZM1 + 1
							else:
								WZM2 = WZM1

							if WZM2 >= 60:
								WZJ = WZJ + 1
								WZM2 = WZM2 - 60
							else:
								WZJ = WZJ
								WZM2 = WZM2

							#Waktu Ashar
							AA = acot(KA + tan(radians(abs(Delta - L))))
							AA1 = degrees(AA)
							HAA = acos((sin(radians(AA1)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAA1 = degrees(HAA)
							WA = Transit + (HAA1/15)
							WAJ = int(WA)
							WA2 = WA - WAJ
							WAM = int(WA2*60)
							WA3 = ((WA2*60) - WAM)
							WAS = int(WA3*60)
							if WAS > 0:
								WAM1 = WAM + 1
							else:
								WAM1 = WAM

							if WAM1 >= 60:
								WAJ = WAJ + 1
								WAM1 = WAM1 - 60
							else:
								WAJ = WAJ
								WAM1 = WAM1

							# Waktu Maghrib
							AM = (-0.8333) - 0.0347*sqrt(H)
							HAM = acos((sin(radians(AM)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAM1 = degrees(HAM)
							WM = Transit + (HAM1/15)
							WMJ = int(WM)
							WM2 = WM - WMJ
							WMM = int(WM2*60)
							WM3 = ((WM2*60) - WMM)
							WMS = int(WM3*60)
							if WMS > 0:
								WMM1 = WMM + 1
							else:
								WMM1 = WMM

							if WMM1 >= 60:
								WMJ = WMJ + 1
								WMM1 = WMM1 - 60
							else:
								WMJ = WMJ
								WMM1 = WMM1

							# Waktu Isya'
							HAI = acos((sin(radians(AI)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAI1 = degrees(HAI)
							WI = Transit + (HAI1/15)
							WIJ = int(WI)
							WI2 = WI - WIJ
							WIM = int(WI2*60)
							WI3 = ((WI2*60) - WIM)
							WIS = int(WI3*60)
							if WIS > 0:
								WIM1 = WIM + 1
							else:
								WIM1 = WIM

							if WIM1 >= 60:
								WIJ = WIJ + 1
								WIM1 = WIM1 - 60
							else:
								WIJ = WIJ
								WIM1 = WIM1

							# Waktu Subuh
							HAS = acos((sin(radians(AS)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAS1 = degrees(HAS)
							WS = Transit - (HAS1/15)
							WSJ = int(WS)
							WS2 = WS - WSJ
							WSM = int(WS2*60)
							WS3 = ((WS2*60) - WSM)
							WSS = int(WS3*60)
							if WSS > 0:
								WSM1 = WSM + 1
							else:
								WSM1 = WSM

							if WSM1 >= 60:
								WSJ = WSJ + 1
								WSM1 = WSM1 - 60
							else:
								WSJ = WSJ
								WSM1 = WSM1

							# Waktu Terbit Matahari
							AT = (-0.8333) - 0.0347*sqrt(H)
							HAT = acos((sin(radians(AT)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAT1 = degrees(HAT)
							WT = Transit - (HAT1/15)
							WTJ = int(WT)
							WT2 = WT - WTJ
							WTM = int(WT2*60)
							WT3 = ((WT2*60) - WTM)
							WTS = int(WT3*60)
							if WTS > 0:
								WTM1 = WTM + 1
							else:
								WTM1 = WTM

							if WTM1 >= 60:
								WTJ = WTJ + 1
								WTM1 = WTM1 - 60
							else:
								WTJ = WTJ
								WTM1 = WTM1

							if Z == 7:
								(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1) = satubulanWIB(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)
							elif Z == 8:
								(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1) = satubulanWITA(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)
							elif Z ==9:
								(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1) = satubulanWIT(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)
					elif Bln == 14:
						for D in range(1,29):
							JD = 1720994.5 + int(365.25 * Th) + int(30.6001 * (Bln + 1)) + B + D
							JD1 = JD - Z/24
							T = 2*3.14*(JD1 - 2451545)/365.25
							Delta = 0.3787 + 23.64*sin(radians(57.297*T - 79.547)) + 0.3812*sin(radians(2*57.297*T - 82.682)) + 0.17132*sin(radians(3*57.297*T - 59.722))
							U = (JD1 - 2451545)/36525
							L0 = (280.46607+36000.7698*U)

							# Membuat L0 Menjadi Sudut Antara 0 - 360
							if L0 < 360:
								L0 = L0
							elif L0 < 720:
								L0 = L0 - 360
							elif L0 < 1081:
								L0 = L0 - 720
							elif L0 < 1441:
								L0 = L0 - 1080
							elif L0 < 1801:
								L0 = L0 - 1440
							elif L0 < 2161:
								L0 = L0 - 1800
							elif L0 < 2521:
								L0 = L0 - 2160
							elif L0 < 2881:
								L0 = L0 - 2520
							elif L0 < 3241:
								L0 = L0 - 2880
							elif L0 < 3601:
								L0 = L0 - 3240
							elif L0 < 3960:
								L0 = L0 - 3600

							# Waktu Transit
							ET = (-1*(1789+237*U)*sin(radians(L0))-(7416-62*U)*cos(radians(L0))+(9934-14*U)*sin(radians(2*L0))-(29+5*U)*cos(radians(2*L0))+(74+10*U)*sin(radians(3*L0))+(320-4*U)*cos(radians(3*L0))-212*sin(radians(4*L0)))/1000
							Transit = 12 + Z - (Bjr/15) - (ET/60)

							# Waktu Zuhur
							WZJ = int(Transit)
							WZ2 = Transit - WZJ
							WZM = (int(WZ2*60))
							WZ3 = ((WZ2*60) - WZM)
							WZS = int(WZ3*60)
							WZM1 = WZM + 2
							if WZS > 0:
								WZM2 = WZM1 + 1
							else:
								WZM2 = WZM1

							if WZM2 >= 60:
								WZJ = WZJ + 1
								WZM2 = WZM2 - 60
							else:
								WZJ = WZJ
								WZM2 = WZM2

							#Waktu Ashar
							AA = acot(KA + tan(radians(abs(Delta - L))))
							AA1 = degrees(AA)
							HAA = acos((sin(radians(AA1)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAA1 = degrees(HAA)
							WA = Transit + (HAA1/15)
							WAJ = int(WA)
							WA2 = WA - WAJ
							WAM = int(WA2*60)
							WA3 = ((WA2*60) - WAM)
							WAS = int(WA3*60)
							if WAS > 0:
								WAM1 = WAM + 1
							else:
								WAM1 = WAM

							if WAM1 >= 60:
								WAJ = WAJ + 1
								WAM1 = WAM1 - 60
							else:
								WAJ = WAJ
								WAM1 = WAM1

							# Waktu Maghrib
							AM = (-0.8333) - 0.0347*sqrt(H)
							HAM = acos((sin(radians(AM)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAM1 = degrees(HAM)
							WM = Transit + (HAM1/15)
							WMJ = int(WM)
							WM2 = WM - WMJ
							WMM = int(WM2*60)
							WM3 = ((WM2*60) - WMM)
							WMS = int(WM3*60)
							if WMS > 0:
								WMM1 = WMM + 1
							else:
								WMM1 = WMM

							if WMM1 >= 60:
								WMJ = WMJ + 1
								WMM1 = WMM1 - 60
							else:
								WMJ = WMJ
								WMM1 = WMM1

							# Waktu Isya'
							HAI = acos((sin(radians(AI)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAI1 = degrees(HAI)
							WI = Transit + (HAI1/15)
							WIJ = int(WI)
							WI2 = WI - WIJ
							WIM = int(WI2*60)
							WI3 = ((WI2*60) - WIM)
							WIS = int(WI3*60)
							if WIS > 0:
								WIM1 = WIM + 1
							else:
								WIM1 = WIM

							if WIM1 >= 60:
								WIJ = WIJ + 1
								WIM1 = WIM1 - 60
							else:
								WIJ = WIJ
								WIM1 = WIM1

							# Waktu Subuh
							HAS = acos((sin(radians(AS)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAS1 = degrees(HAS)
							WS = Transit - (HAS1/15)
							WSJ = int(WS)
							WS2 = WS - WSJ
							WSM = int(WS2*60)
							WS3 = ((WS2*60) - WSM)
							WSS = int(WS3*60)
							if WSS > 0:
								WSM1 = WSM + 1
							else:
								WSM1 = WSM

							if WSM1 >= 60:
								WSJ = WSJ + 1
								WSM1 = WSM1 - 60
							else:
								WSJ = WSJ
								WSM1 = WSM1

							# Waktu Terbit Matahari
							AT = (-0.8333) - 0.0347*sqrt(H)
							HAT = acos((sin(radians(AT)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAT1 = degrees(HAT)
							WT = Transit - (HAT1/15)
							WTJ = int(WT)
							WT2 = WT - WTJ
							WTM = int(WT2*60)
							WT3 = ((WT2*60) - WTM)
							WTS = int(WT3*60)
							if WTS > 0:
								WTM1 = WTM + 1
							else:
								WTM1 = WTM

							if WTM1 >= 60:
								WTJ = WTJ + 1
								WTM1 = WTM1 - 60
							else:
								WTJ = WTJ
								WTM1 = WTM1

							if Z == 7:
								(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1) = satubulanWIB(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)
							elif Z == 8:
								(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1) = satubulanWITA(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)
							elif Z ==9:
								(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1) = satubulanWIT(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)
					else:
						for D in range(1,31):
							JD = 1720994.5 + int(365.25 * Th) + int(30.6001 * (Bln + 1)) + B + D
							JD1 = JD - Z/24
							T = 2*3.14*(JD1 - 2451545)/365.25
							Delta = 0.3787 + 23.64*sin(radians(57.297*T - 79.547)) + 0.3812*sin(radians(2*57.297*T - 82.682)) + 0.17132*sin(radians(3*57.297*T - 59.722))
							U = (JD1 - 2451545)/36525
							L0 = (280.46607+36000.7698*U)

							# Membuat L0 Menjadi Sudut Antara 0 - 360
							if L0 < 360:
								L0 = L0
							elif L0 < 720:
								L0 = L0 - 360
							elif L0 < 1081:
								L0 = L0 - 720
							elif L0 < 1441:
								L0 = L0 - 1080
							elif L0 < 1801:
								L0 = L0 - 1440
							elif L0 < 2161:
								L0 = L0 - 1800
							elif L0 < 2521:
								L0 = L0 - 2160
							elif L0 < 2881:
								L0 = L0 - 2520
							elif L0 < 3241:
								L0 = L0 - 2880
							elif L0 < 3601:
								L0 = L0 - 3240
							elif L0 < 3960:
								L0 = L0 - 3600

							# Waktu Transit
							ET = (-1*(1789+237*U)*sin(radians(L0))-(7416-62*U)*cos(radians(L0))+(9934-14*U)*sin(radians(2*L0))-(29+5*U)*cos(radians(2*L0))+(74+10*U)*sin(radians(3*L0))+(320-4*U)*cos(radians(3*L0))-212*sin(radians(4*L0)))/1000
							Transit = 12 + Z - (Bjr/15) - (ET/60)

							# Waktu Zuhur
							WZJ = int(Transit)
							WZ2 = Transit - WZJ
							WZM = (int(WZ2*60))
							WZ3 = ((WZ2*60) - WZM)
							WZS = int(WZ3*60)
							WZM1 = WZM + 2
							if WZS > 0:
								WZM2 = WZM1 + 1
							else:
								WZM2 = WZM1

							if WZM2 >= 60:
								WZJ = WZJ + 1
								WZM2 = WZM2 - 60
							else:
								WZJ = WZJ
								WZM2 = WZM2

							#Waktu Ashar
							AA = acot(KA + tan(radians(abs(Delta - L))))
							AA1 = degrees(AA)
							HAA = acos((sin(radians(AA1)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAA1 = degrees(HAA)
							WA = Transit + (HAA1/15)
							WAJ = int(WA)
							WA2 = WA - WAJ
							WAM = int(WA2*60)
							WA3 = ((WA2*60) - WAM)
							WAS = int(WA3*60)
							if WAS > 0:
								WAM1 = WAM + 1
							else:
								WAM1 = WAM

							if WAM1 >= 60:
								WAJ = WAJ + 1
								WAM1 = WAM1 - 60
							else:
								WAJ = WAJ
								WAM1 = WAM1

							# Waktu Maghrib
							AM = (-0.8333) - 0.0347*sqrt(H)
							HAM = acos((sin(radians(AM)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAM1 = degrees(HAM)
							WM = Transit + (HAM1/15)
							WMJ = int(WM)
							WM2 = WM - WMJ
							WMM = int(WM2*60)
							WM3 = ((WM2*60) - WMM)
							WMS = int(WM3*60)
							if WMS > 0:
								WMM1 = WMM + 1
							else:
								WMM1 = WMM

							if WMM1 >= 60:
								WMJ = WMJ + 1
								WMM1 = WMM1 - 60
							else:
								WMJ = WMJ
								WMM1 = WMM1

							# Waktu Isya'
							HAI = acos((sin(radians(AI)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAI1 = degrees(HAI)
							WI = Transit + (HAI1/15)
							WIJ = int(WI)
							WI2 = WI - WIJ
							WIM = int(WI2*60)
							WI3 = ((WI2*60) - WIM)
							WIS = int(WI3*60)
							if WIS > 0:
								WIM1 = WIM + 1
							else:
								WIM1 = WIM

							if WIM1 >= 60:
								WIJ = WIJ + 1
								WIM1 = WIM1 - 60
							else:
								WIJ = WIJ
								WIM1 = WIM1

							# Waktu Subuh
							HAS = acos((sin(radians(AS)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAS1 = degrees(HAS)
							WS = Transit - (HAS1/15)
							WSJ = int(WS)
							WS2 = WS - WSJ
							WSM = int(WS2*60)
							WS3 = ((WS2*60) - WSM)
							WSS = int(WS3*60)
							if WSS > 0:
								WSM1 = WSM + 1
							else:
								WSM1 = WSM

							if WSM1 >= 60:
								WSJ = WSJ + 1
								WSM1 = WSM1 - 60
							else:
								WSJ = WSJ
								WSM1 = WSM1

							# Waktu Terbit Matahari
							AT = (-0.8333) - 0.0347*sqrt(H)
							HAT = acos((sin(radians(AT)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAT1 = degrees(HAT)
							WT = Transit - (HAT1/15)
							WTJ = int(WT)
							WT2 = WT - WTJ
							WTM = int(WT2*60)
							WT3 = ((WT2*60) - WTM)
							WTS = int(WT3*60)
							if WTS > 0:
								WTM1 = WTM + 1
							else:
								WTM1 = WTM

							if WTM1 >= 60:
								WTJ = WTJ + 1
								WTM1 = WTM1 - 60
							else:
								WTJ = WTJ
								WTM1 = WTM1

							if Z == 7:
								(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1) = satubulanWIB(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)
							elif Z == 8:
								(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1) = satubulanWITA(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)
							elif Z ==9:
								(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1) = satubulanWIT(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)
			
			# Perhitungan 1 Bulan Hijriah
			elif M == 'Muharram' or M == 'Safar' or M == 'Rabiul Awwal' or M == 'Rabiul Akhir' or M == 'Jumadil Awwal' or M == 'Jumadil Akhir' or M == 'Rajab' or M == 'Syaban' or M == 'Ramadhan' or M == 'Syawal' or M == 'Dzulqaidah' or M == 'Dzulhijjah':
				# 1 Bulan Kabisat
				if Y % 30 == 2 or Y % 30 == 5 or Y % 30 == 7 or Y % 30 == 10 or Y % 30 == 13 or Y % 30 == 16 or Y % 30 == 18 or Y % 30 == 21 or Y % 30 == 24 or Y % 30 == 26 or Y % 30 == 29:
					if M1 == 1 or M1 == 3 or M1 == 5 or M1 == 7 or M1 == 11:
						for D in range(1,30):
							D4 = D+D1+D2+D3
							JD = 1948438.5+D4
							JD1 = JD - Z/24
							T = 2*3.14*(JD1 - 2451545)/365.25
							Delta = 0.3787 + 23.64*sin(radians(57.297*T - 79.547)) + 0.3812*sin(radians(2*57.297*T - 82.682)) + 0.17132*sin(radians(3*57.297*T - 59.722))
							U = (JD1 - 2451545)/36525
							L0 = (280.46607+36000.7698*U)
							if L0 < 360:
								L0 = L0
							elif L0 < 720:
								L0 = L0 - 360
							elif L0 < 1081:
								L0 = L0 - 720
							elif L0 < 1441:
								L0 = L0 - 1080
							elif L0 < 1801:
								L0 = L0 - 1440
							elif L0 < 2161:
								L0 = L0 - 1800
							elif L0 < 2521:
								L0 = L0 - 2160
							elif L0 < 2881:
								L0 = L0 - 2520
							elif L0 < 3241:
								L0 = L0 - 2880
							elif L0 < 3601:
								L0 = L0 - 3240
							elif L0 < 3960:
								L0 = L0 - 3600

							# Waktu Transit
							ET = (-1*(1789+237*U)*sin(radians(L0))-(7416-62*U)*cos(radians(L0))+(9934-14*U)*sin(radians(2*L0))-(29+5*U)*cos(radians(2*L0))+(74+10*U)*sin(radians(3*L0))+(320-4*U)*cos(radians(3*L0))-212*sin(radians(4*L0)))/1000
							Transit = 12 + Z - (Bjr/15) - (ET/60)

							# Waktu Zuhur
							WZJ = int(Transit)
							WZ2 = Transit - WZJ
							WZM = (int(WZ2*60))
							WZ3 = ((WZ2*60) - WZM)
							WZS = int(WZ3*60)
							WZM1 = WZM + 2
							if WZS > 0:
								WZM2 = WZM1 + 1
							else:
								WZM2 = WZM1

							if WZM2 >= 60:
								WZJ = WZJ + 1
								WZM2 = WZM2 - 60
							else:
								WZJ = WZJ
								WZM2 = WZM2

							#Waktu Ashar
							AA = acot(KA + tan(radians(abs(Delta - L))))
							AA1 = degrees(AA)
							HAA = acos((sin(radians(AA1)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAA1 = degrees(HAA)
							WA = Transit + (HAA1/15)
							WAJ = int(WA)
							WA2 = WA - WAJ
							WAM = int(WA2*60)
							WA3 = ((WA2*60) - WAM)
							WAS = int(WA3*60)
							if WAS > 0:
								WAM1 = WAM + 1
							else:
								WAM1 = WAM

							if WAM1 >= 60:
								WAJ = WAJ + 1
								WAM1 = WAM1 - 60
							else:
								WAJ = WAJ
								WAM1 = WAM1

							# Waktu Maghrib
							AM = (-0.8333) - 0.0347*sqrt(H)
							HAM = acos((sin(radians(AM)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAM1 = degrees(HAM)
							WM = Transit + (HAM1/15)
							WMJ = int(WM)
							WM2 = WM - WMJ
							WMM = int(WM2*60)
							WM3 = ((WM2*60) - WMM)
							WMS = int(WM3*60)
							if WMS > 0:
								WMM1 = WMM + 1
							else:
								WMM1 = WMM

							if WMM1 >= 60:
								WMJ = WMJ + 1
								WMM1 = WMM1 - 60
							else:
								WMJ = WMJ
								WMM1 = WMM1

							# Waktu Isya'
							HAI = acos((sin(radians(AI)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAI1 = degrees(HAI)
							WI = Transit + (HAI1/15)
							WIJ = int(WI)
							WI2 = WI - WIJ
							WIM = int(WI2*60)
							WI3 = ((WI2*60) - WIM)
							WIS = int(WI3*60)
							if WIS > 0:
								WIM1 = WIM + 1
							else:
								WIM1 = WIM

							if WIM1 >= 60:
								WIJ = WIJ + 1
								WIM1 = WIM1 - 60
							else:
								WIJ = WIJ
								WIM1 = WIM1

							# Waktu Subuh
							HAS = acos((sin(radians(AS)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAS1 = degrees(HAS)
							WS = Transit - (HAS1/15)
							WSJ = int(WS)
							WS2 = WS - WSJ
							WSM = int(WS2*60)
							WS3 = ((WS2*60) - WSM)
							WSS = int(WS3*60)
							if WSS > 0:
								WSM1 = WSM + 1
							else:
								WSM1 = WSM

							if WSM1 >= 60:
								WSJ = WSJ + 1
								WSM1 = WSM1 - 60
							else:
								WSJ = WSJ
								WSM1 = WSM1

							# Waktu Terbit Matahari
							AT = (-0.8333) - 0.0347*sqrt(H)
							HAT = acos((sin(radians(AT)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAT1 = degrees(HAT)
							WT = Transit - (HAT1/15)
							WTJ = int(WT)
							WT2 = WT - WTJ
							WTM = int(WT2*60)
							WT3 = ((WT2*60) - WTM)
							WTS = int(WT3*60)
							if WTS > 0:
								WTM1 = WTM + 1
							else:
								WTM1 = WTM

							if WTM1 >= 60:
								WTJ = WTJ + 1
								WTM1 = WTM1 - 60
							else:
								WTJ = WTJ
								WTM1 = WTM1

							if Z == 7:
								(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1) = satubulanWIB(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)
							elif Z == 8:
								(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1) = satubulanWITA(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)
							elif Z ==9:
								(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1) = satubulanWIT(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)
					else:
						for D in range(1,31):
							D4 = D+D1+D2+D3
							JD = 1948438.5+D4
							JD1 = JD - Z/24
							T = 2*3.14*(JD1 - 2451545)/365.25
							Delta = 0.3787 + 23.64*sin(radians(57.297*T - 79.547)) + 0.3812*sin(radians(2*57.297*T - 82.682)) + 0.17132*sin(radians(3*57.297*T - 59.722))
							U = (JD1 - 2451545)/36525
							L0 = (280.46607+36000.7698*U)

							# Membuat L0 Menjadi Sudut Antara 0 - 360
							if L0 < 360:
								L0 = L0
							elif L0 < 720:
								L0 = L0 - 360
							elif L0 < 1081:
								L0 = L0 - 720
							elif L0 < 1441:
								L0 = L0 - 1080
							elif L0 < 1801:
								L0 = L0 - 1440
							elif L0 < 2161:
								L0 = L0 - 1800
							elif L0 < 2521:
								L0 = L0 - 2160
							elif L0 < 2881:
								L0 = L0 - 2520
							elif L0 < 3241:
								L0 = L0 - 2880
							elif L0 < 3601:
								L0 = L0 - 3240
							elif L0 < 3960:
								L0 = L0 - 3600

							# Waktu Transit
							ET = (-1*(1789+237*U)*sin(radians(L0))-(7416-62*U)*cos(radians(L0))+(9934-14*U)*sin(radians(2*L0))-(29+5*U)*cos(radians(2*L0))+(74+10*U)*sin(radians(3*L0))+(320-4*U)*cos(radians(3*L0))-212*sin(radians(4*L0)))/1000
							Transit = 12 + Z - (Bjr/15) - (ET/60)

							# Waktu Zuhur
							WZJ = int(Transit)
							WZ2 = Transit - WZJ
							WZM = (int(WZ2*60))
							WZ3 = ((WZ2*60) - WZM)
							WZS = int(WZ3*60)
							WZM1 = WZM + 2
							if WZS > 0:
								WZM2 = WZM1 + 1
							else:
								WZM2 = WZM1

							if WZM2 >= 60:
								WZJ = WZJ + 1
								WZM2 = WZM2 - 60
							else:
								WZJ = WZJ
								WZM2 = WZM2

							#Waktu Ashar
							AA = acot(KA + tan(radians(abs(Delta - L))))
							AA1 = degrees(AA)
							HAA = acos((sin(radians(AA1)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAA1 = degrees(HAA)
							WA = Transit + (HAA1/15)
							WAJ = int(WA)
							WA2 = WA - WAJ
							WAM = int(WA2*60)
							WA3 = ((WA2*60) - WAM)
							WAS = int(WA3*60)
							if WAS > 0:
								WAM1 = WAM + 1
							else:
								WAM1 = WAM

							if WAM1 >= 60:
								WAJ = WAJ + 1
								WAM1 = WAM1 - 60
							else:
								WAJ = WAJ
								WAM1 = WAM1

							# Waktu Maghrib
							AM = (-0.8333) - 0.0347*sqrt(H)
							HAM = acos((sin(radians(AM)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAM1 = degrees(HAM)
							WM = Transit + (HAM1/15)
							WMJ = int(WM)
							WM2 = WM - WMJ
							WMM = int(WM2*60)
							WM3 = ((WM2*60) - WMM)
							WMS = int(WM3*60)
							if WMS > 0:
								WMM1 = WMM + 1
							else:
								WMM1 = WMM

							if WMM1 >= 60:
								WMJ = WMJ + 1
								WMM1 = WMM1 - 60
							else:
								WMJ = WMJ
								WMM1 = WMM1

							# Waktu Isya'
							HAI = acos((sin(radians(AI)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAI1 = degrees(HAI)
							WI = Transit + (HAI1/15)
							WIJ = int(WI)
							WI2 = WI - WIJ
							WIM = int(WI2*60)
							WI3 = ((WI2*60) - WIM)
							WIS = int(WI3*60)
							if WIS > 0:
								WIM1 = WIM + 1
							else:
								WIM1 = WIM

							if WIM1 >= 60:
								WIJ = WIJ + 1
								WIM1 = WIM1 - 60
							else:
								WIJ = WIJ
								WIM1 = WIM1

							# Waktu Subuh
							HAS = acos((sin(radians(AS)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAS1 = degrees(HAS)
							WS = Transit - (HAS1/15)
							WSJ = int(WS)
							WS2 = WS - WSJ
							WSM = int(WS2*60)
							WS3 = ((WS2*60) - WSM)
							WSS = int(WS3*60)
							if WSS > 0:
								WSM1 = WSM + 1
							else:
								WSM1 = WSM

							if WSM1 >= 60:
								WSJ = WSJ + 1
								WSM1 = WSM1 - 60
							else:
								WSJ = WSJ
								WSM1 = WSM1

							# Waktu Terbit Matahari
							AT = (-0.8333) - 0.0347*sqrt(H)
							HAT = acos((sin(radians(AT)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAT1 = degrees(HAT)
							WT = Transit - (HAT1/15)
							WTJ = int(WT)
							WT2 = WT - WTJ
							WTM = int(WT2*60)
							WT3 = ((WT2*60) - WTM)
							WTS = int(WT3*60)
							if WTS > 0:
								WTM1 = WTM + 1
							else:
								WTM1 = WTM

							if WTM1 >= 60:
								WTJ = WTJ + 1
								WTM1 = WTM1 - 60
							else:
								WTJ = WTJ
								WTM1 = WTM1

							if Z == 7:
								(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1) = satubulanWIB(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)
							elif Z == 8:
								(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1) = satubulanWITA(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)
							elif Z ==9:
								(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1) = satubulanWIT(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)
				else:
					if M1 == 1 or M1 == 3 or M1 == 5 or M1 == 7 or M1 == 11 or M1 == 12:
						for D in range(1,30):
							D4 = D+D1+D2+D3
							JD = 1948438.5+D4
							JD1 = JD - Z/24
							T = 2*3.14*(JD1 - 2451545)/365.25
							Delta = 0.3787 + 23.64*sin(radians(57.297*T - 79.547)) + 0.3812*sin(radians(2*57.297*T - 82.682)) + 0.17132*sin(radians(3*57.297*T - 59.722))
							U = (JD1 - 2451545)/36525
							L0 = (280.46607+36000.7698*U)

							# Membuat L0 Menjadi Sudut Antara 0 - 360
							if L0 < 360:
								L0 = L0
							elif L0 < 720:
								L0 = L0 - 360
							elif L0 < 1081:
								L0 = L0 - 720
							elif L0 < 1441:
								L0 = L0 - 1080
							elif L0 < 1801:
								L0 = L0 - 1440
							elif L0 < 2161:
								L0 = L0 - 1800
							elif L0 < 2521:
								L0 = L0 - 2160
							elif L0 < 2881:
								L0 = L0 - 2520
							elif L0 < 3241:
								L0 = L0 - 2880
							elif L0 < 3601:
								L0 = L0 - 3240
							elif L0 < 3960:
								L0 = L0 - 3600

							# Waktu Transit
							ET = (-1*(1789+237*U)*sin(radians(L0))-(7416-62*U)*cos(radians(L0))+(9934-14*U)*sin(radians(2*L0))-(29+5*U)*cos(radians(2*L0))+(74+10*U)*sin(radians(3*L0))+(320-4*U)*cos(radians(3*L0))-212*sin(radians(4*L0)))/1000
							Transit = 12 + Z - (Bjr/15) - (ET/60)

							# Waktu Zuhur
							WZJ = int(Transit)
							WZ2 = Transit - WZJ
							WZM = (int(WZ2*60))
							WZ3 = ((WZ2*60) - WZM)
							WZS = int(WZ3*60)
							WZM1 = WZM + 2
							if WZS > 0:
								WZM2 = WZM1 + 1
							else:
								WZM2 = WZM1

							if WZM2 >= 60:
								WZJ = WZJ + 1
								WZM2 = WZM2 - 60
							else:
								WZJ = WZJ
								WZM2 = WZM2

							#Waktu Ashar
							AA = acot(KA + tan(radians(abs(Delta - L))))
							AA1 = degrees(AA)
							HAA = acos((sin(radians(AA1)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAA1 = degrees(HAA)
							WA = Transit + (HAA1/15)
							WAJ = int(WA)
							WA2 = WA - WAJ
							WAM = int(WA2*60)
							WA3 = ((WA2*60) - WAM)
							WAS = int(WA3*60)
							if WAS > 0:
								WAM1 = WAM + 1
							else:
								WAM1 = WAM

							if WAM1 >= 60:
								WAJ = WAJ + 1
								WAM1 = WAM1 - 60
							else:
								WAJ = WAJ
								WAM1 = WAM1

							# Waktu Maghrib
							AM = (-0.8333) - 0.0347*sqrt(H)
							HAM = acos((sin(radians(AM)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAM1 = degrees(HAM)
							WM = Transit + (HAM1/15)
							WMJ = int(WM)
							WM2 = WM - WMJ
							WMM = int(WM2*60)
							WM3 = ((WM2*60) - WMM)
							WMS = int(WM3*60)
							if WMS > 0:
								WMM1 = WMM + 1
							else:
								WMM1 = WMM

							if WMM1 >= 60:
								WMJ = WMJ + 1
								WMM1 = WMM1 - 60
							else:
								WMJ = WMJ
								WMM1 = WMM1

							# Waktu Isya'
							HAI = acos((sin(radians(AI)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAI1 = degrees(HAI)
							WI = Transit + (HAI1/15)
							WIJ = int(WI)
							WI2 = WI - WIJ
							WIM = int(WI2*60)
							WI3 = ((WI2*60) - WIM)
							WIS = int(WI3*60)
							if WIS > 0:
								WIM1 = WIM + 1
							else:
								WIM1 = WIM

							if WIM1 >= 60:
								WIJ = WIJ + 1
								WIM1 = WIM1 - 60
							else:
								WIJ = WIJ
								WIM1 = WIM1

							# Waktu Subuh
							HAS = acos((sin(radians(AS)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAS1 = degrees(HAS)
							WS = Transit - (HAS1/15)
							WSJ = int(WS)
							WS2 = WS - WSJ
							WSM = int(WS2*60)
							WS3 = ((WS2*60) - WSM)
							WSS = int(WS3*60)
							if WSS > 0:
								WSM1 = WSM + 1
							else:
								WSM1 = WSM

							if WSM1 >= 60:
								WSJ = WSJ + 1
								WSM1 = WSM1 - 60
							else:
								WSJ = WSJ
								WSM1 = WSM1

							# Waktu Terbit Matahari
							AT = (-0.8333) - 0.0347*sqrt(H)
							HAT = acos((sin(radians(AT)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAT1 = degrees(HAT)
							WT = Transit - (HAT1/15)
							WTJ = int(WT)
							WT2 = WT - WTJ
							WTM = int(WT2*60)
							WT3 = ((WT2*60) - WTM)
							WTS = int(WT3*60)
							if WTS > 0:
								WTM1 = WTM + 1
							else:
								WTM1 = WTM

							if WTM1 >= 60:
								WTJ = WTJ + 1
								WTM1 = WTM1 - 60
							else:
								WTJ = WTJ
								WTM1 = WTM1

							if Z == 7:
								(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1) = satubulanWIB(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)
							elif Z == 8:
								(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1) = satubulanWITA(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)
							elif Z ==9:
								(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1) = satubulanWIT(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)
					else:
						for D in range(1,31):
							D4 = D+D1+D2+D3
							JD = 1948438.5+D4
							JD1 = JD - Z/24
							T = 2*3.14*(JD1 - 2451545)/365.25
							Delta = 0.3787 + 23.64*sin(radians(57.297*T - 79.547)) + 0.3812*sin(radians(2*57.297*T - 82.682)) + 0.17132*sin(radians(3*57.297*T - 59.722))
							U = (JD1 - 2451545)/36525
							L0 = (280.46607+36000.7698*U)

							# Membuat L0 Menjadi Sudut Antara 0 - 360
							if L0 < 360:
								L0 = L0
							elif L0 < 720:
								L0 = L0 - 360
							elif L0 < 1081:
								L0 = L0 - 720
							elif L0 < 1441:
								L0 = L0 - 1080
							elif L0 < 1801:
								L0 = L0 - 1440
							elif L0 < 2161:
								L0 = L0 - 1800
							elif L0 < 2521:
								L0 = L0 - 2160
							elif L0 < 2881:
								L0 = L0 - 2520
							elif L0 < 3241:
								L0 = L0 - 2880
							elif L0 < 3601:
								L0 = L0 - 3240
							elif L0 < 3960:
								L0 = L0 - 3600

							# Waktu Transit
							ET = (-1*(1789+237*U)*sin(radians(L0))-(7416-62*U)*cos(radians(L0))+(9934-14*U)*sin(radians(2*L0))-(29+5*U)*cos(radians(2*L0))+(74+10*U)*sin(radians(3*L0))+(320-4*U)*cos(radians(3*L0))-212*sin(radians(4*L0)))/1000
							Transit = 12 + Z - (Bjr/15) - (ET/60)

							# Waktu Zuhur
							WZJ = int(Transit)
							WZ2 = Transit - WZJ
							WZM = (int(WZ2*60))
							WZ3 = ((WZ2*60) - WZM)
							WZS = int(WZ3*60)
							WZM1 = WZM + 2
							if WZS > 0:
								WZM2 = WZM1 + 1
							else:
								WZM2 = WZM1

							if WZM2 >= 60:
								WZJ = WZJ + 1
								WZM2 = WZM2 - 60
							else:
								WZJ = WZJ
								WZM2 = WZM2

							#Waktu Ashar
							AA = acot(KA + tan(radians(abs(Delta - L))))
							AA1 = degrees(AA)
							HAA = acos((sin(radians(AA1)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAA1 = degrees(HAA)
							WA = Transit + (HAA1/15)
							WAJ = int(WA)
							WA2 = WA - WAJ
							WAM = int(WA2*60)
							WA3 = ((WA2*60) - WAM)
							WAS = int(WA3*60)
							if WAS > 0:
								WAM1 = WAM + 1
							else:
								WAM1 = WAM

							if WAM1 >= 60:
								WAJ = WAJ + 1
								WAM1 = WAM1 - 60
							else:
								WAJ = WAJ
								WAM1 = WAM1

							# Waktu Maghrib
							AM = (-0.8333) - 0.0347*sqrt(H)
							HAM = acos((sin(radians(AM)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAM1 = degrees(HAM)
							WM = Transit + (HAM1/15)
							WMJ = int(WM)
							WM2 = WM - WMJ
							WMM = int(WM2*60)
							WM3 = ((WM2*60) - WMM)
							WMS = int(WM3*60)
							if WMS > 0:
								WMM1 = WMM + 1
							else:
								WMM1 = WMM

							if WMM1 >= 60:
								WMJ = WMJ + 1
								WMM1 = WMM1 - 60
							else:
								WMJ = WMJ
								WMM1 = WMM1

							# Waktu Isya'
							HAI = acos((sin(radians(AI)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAI1 = degrees(HAI)
							WI = Transit + (HAI1/15)
							WIJ = int(WI)
							WI2 = WI - WIJ
							WIM = int(WI2*60)
							WI3 = ((WI2*60) - WIM)
							WIS = int(WI3*60)
							if WIS > 0:
								WIM1 = WIM + 1
							else:
								WIM1 = WIM

							if WIM1 >= 60:
								WIJ = WIJ + 1
								WIM1 = WIM1 - 60
							else:
								WIJ = WIJ
								WIM1 = WIM1

							# Waktu Subuh
							HAS = acos((sin(radians(AS)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAS1 = degrees(HAS)
							WS = Transit - (HAS1/15)
							WSJ = int(WS)
							WS2 = WS - WSJ
							WSM = int(WS2*60)
							WS3 = ((WS2*60) - WSM)
							WSS = int(WS3*60)
							if WSS > 0:
								WSM1 = WSM + 1
							else:
								WSM1 = WSM

							if WSM1 >= 60:
								WSJ = WSJ + 1
								WSM1 = WSM1 - 60
							else:
								WSJ = WSJ
								WSM1 = WSM1

							# Waktu Terbit Matahari
							AT = (-0.8333) - 0.0347*sqrt(H)
							HAT = acos((sin(radians(AT)) - sin(radians(L))*sin(radians(Delta)))/(cos(radians(L))*cos(radians(Delta))))
							HAT1 = degrees(HAT)
							WT = Transit - (HAT1/15)
							WTJ = int(WT)
							WT2 = WT - WTJ
							WTM = int(WT2*60)
							WT3 = ((WT2*60) - WTM)
							WTS = int(WT3*60)
							if WTS > 0:
								WTM1 = WTM + 1
							else:
								WTM1 = WTM

							if WTM1 >= 60:
								WTJ = WTJ + 1
								WTM1 = WTM1 - 60
							else:
								WTJ = WTJ
								WTM1 = WTM1

							if Z == 7:
								(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1) = satubulanWIB(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)
							elif Z == 8:
								(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1) = satubulanWITA(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)
							elif Z ==9:
								(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1) = satubulanWIT(WSJ,WSM1,WTJ,WTM1,WZJ,WZM2,WAJ,WAM1,WMJ,WMM1,WIJ,WIM1)

	ditampilkan = waktu_shalat(tanggal.get(), bulan.get(), tahun.get(), kota.get())
	ditampilkan.hasil()

tombol = Button(root, text="Cari Waktu Shalat",command=cetak).place(x=80,y=260)

root.mainloop()